"""Vérification import-contacts — import en lot de noms de décideurs via Excel."""
import io
import json
import urllib.request
import urllib.parse
import openpyxl

BASE = "http://localhost:8000/api"
ok, fail = 0, 0

def check(label, cond, detail=""):
    global ok, fail
    if cond:
        ok += 1; print(f"  [OK]   {label}")
    else:
        fail += 1; print(f"  [FAIL] {label} {detail}")

def get(path):
    return json.loads(urllib.request.urlopen(f"{BASE}{path}").read())

def post(path, body):
    data = json.dumps(body).encode()
    req = urllib.request.Request(f"{BASE}{path}", data=data,
        headers={"Content-Type": "application/json"}, method="POST")
    try:
        resp = urllib.request.urlopen(req)
        return resp.status, json.loads(resp.read())
    except urllib.error.HTTPError as e:
        return e.code, json.loads(e.read())

def post_file(path, file_data, filename="test_contacts.xlsx"):
    boundary = "----TestBoundary12345"
    parts = [
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"file\"; filename=\"{filename}\"\r\n\r\n".encode(),
        file_data,
        f"\r\n--{boundary}--\r\n".encode(),
    ]
    data = b"".join(parts)
    req = urllib.request.Request(f"{BASE}{path}", data=data,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"}, method="POST")
    try:
        resp = urllib.request.urlopen(req)
        return resp.status, json.loads(resp.read())
    except urllib.error.HTTPError as e:
        return e.code, json.loads(e.read())

# ── Setup: insert test prospects without contact names ──
print("=== Setup: insert test prospects without contacts ===")
existing = get("/prospects")
test_keywords = ["IC-Test Harnois", "IC-Test Kenworth", "IC-Test Bombardier"]
has_test = any(any(kw in str(p.get("entreprise", "")) for kw in test_keywords) for p in existing)
if not has_test:
    test_prospects = [
        {"entreprise": "IC-Test Harnois Inc.", "contact": "", "adresse": "1000 rue Jarry Est", "ville": "Montréal H1J", "statut": "Courriel envoyé", "telephone": "514-555-1234", "zone": "ZONE 1", "rue": "jarry est", "segment": "Industriel / Commercial", "next_action": "", "contacte": False, "date_contact": "", "notes": "Lettre envoyée"},
        {"entreprise": "IC-Test Kenworth Ltée", "contact": "", "adresse": "2000 boul Métropolitain Est", "ville": "Montréal H1K", "statut": "À contacter / À appeler", "telephone": "514-555-5678", "zone": "ZONE 2", "rue": "metropolitain est", "segment": "Industriel / Commercial", "next_action": "", "contacte": False, "date_contact": "", "notes": ""},
        {"entreprise": "IC-Test Bombardier Corp", "contact": "Jean Déjà", "adresse": "500 rue Bombardier", "ville": "Montréal H1J", "statut": "Courriel envoyé", "telephone": "", "zone": "ZONE 1", "rue": "bombardier", "segment": "Industriel / Commercial", "next_action": "", "contacte": False, "date_contact": "", "notes": ""},
    ]
    all_p = existing + test_prospects
    post("/prospects", {"prospects": all_p})
    print(f"  Inserted {len(test_prospects)} test prospects (2 without contact, 1 with existing)")
else:
    print("  Test prospects already exist")

print("\n=== IC1. Import Excel with decision-maker names ===")
# Build Excel file
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["Entreprise", "Nom_Gestionnaire"])
ws.append(["IC-Test Harnois Inc.", "Jean Harnois"])
ws.append(["IC-Test Kenworth Ltée", "Marc Tremblay"])
ws.append(["IC-Test Bombardier Corp", "Philippe Dubois"])  # should be skipped (already has contact)
ws.append(["Entreprise Inexistante", "Personne"])  # should be not_found
buf = io.BytesIO()
wb.save(buf)
excel_data = buf.getvalue()

status, data = post_file("/prospects/import-contacts", excel_data)
check("Endpoint répond 200", status == 200, f"got {status}: {data}")
check("Status ok", data.get("status") == "ok", f"got {data.get('status')}")
check("2 noms importés (Harnois + Kenworth)", data.get("updated") == 2, f"got {data.get('updated')}")
check("1 ignoré (Bombardier a déjà un contact)", data.get("skipped_existing") == 1, f"got {data.get('skipped_existing')}")
check("1 non trouvé (Entreprise Inexistante)", data.get("not_found") == 1, f"got {data.get('not_found')}")

print("\n=== IC2. Vérification en DB ===")
prospects = get("/prospects")
harnois = next((p for p in prospects if "IC-Test Harnois" in str(p.get("entreprise", ""))), None)
kenworth = next((p for p in prospects if "IC-Test Kenworth" in str(p.get("entreprise", ""))), None)
bombardier = next((p for p in prospects if "IC-Test Bombardier" in str(p.get("entreprise", ""))), None)
check("Harnois → contact = 'Jean Harnois'", harnois and harnois.get("contact") == "Jean Harnois", f"got {harnois.get('contact') if harnois else 'N/A'}")
check("Kenworth → contact = 'Marc Tremblay'", kenworth and kenworth.get("contact") == "Marc Tremblay", f"got {kenworth.get('contact') if kenworth else 'N/A'}")
check("Bombardier → contact inchangé ('Jean Déjà')", bombardier and bombardier.get("contact") == "Jean Déjà", f"got {bombardier.get('contact') if bombardier else 'N/A'}")

print("\n=== IC3. Fuzzy matching (faute de frappe) ===")
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.append(["Entreprise", "Contact"])
ws2.append(["IC-Test Harnois", "Test Fuzzy"])  # missing "Inc." — should still fuzzy match
buf2 = io.BytesIO()
wb2.save(buf2)
# But Harnois already has a contact now, so it should be skipped
status2, data2 = post_file("/prospects/import-contacts", buf2.getvalue())
check("Fuzzy match: Harnois trouvé (ignoré car déjà renseigné)", data2.get("skipped_existing") >= 1, f"got {data2.get('skipped_existing')}")

print("\n=== IC4. Colonnes flexibles ===")
wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.append(["Nom_Syndicat", "Gestionnaire"])
ws3.append(["IC-Test Kenworth", "Nouveau Nom"])  # already has contact → skipped
buf3 = io.BytesIO()
wb3.save(buf3)
status3, data3 = post_file("/prospects/import-contacts", buf3.getvalue())
check("Colonnes alternatives (Nom_Syndicat/Gestionnaire) acceptées", status3 == 200, f"got {status3}")

print("\n=== IC5. Re-import (idempotent) ===")
# Re-import same file — all should be skipped now
status4, data4 = post_file("/prospects/import-contacts", excel_data)
check("Re-import: 0 mis à jour (tous déjà renseignés)", data4.get("updated") == 0, f"got {data4.get('updated')}")
check("Re-import: aucun nouveau nom ajouté", data4.get("updated") == 0 and data4.get("skipped_existing", 0) >= 2, f"got updated={data4.get('updated')} skipped={data4.get('skipped_existing')}")

print("\n=== IC6. Journal d'activité ===")
try:
    logs = get("/activity")
    log_found = any(l.get("action") == "contacts_imported" for l in (logs if isinstance(logs, list) else []))
    check("Action 'contacts_imported' loggée", log_found)
except Exception:
    check("Action 'contacts_imported' loggée", False, "endpoint /activity inaccessible")

print(f"\n{'='*50}\nRESULTAT: {ok} OK / {fail} ECHEC(S)")

# ── Nettoyage ──
all_p = get("/prospects")
kept = [p for p in all_p if not any(kw in str(p.get("entreprise", "")) for kw in test_keywords)]
if len(kept) != len(all_p):
    post("/prospects", {"prospects": kept})
    print(f"Removed {len(all_p) - len(kept)} IC-Test prospects.")
raise SystemExit(1 if fail else 0)
